from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from datetime import datetime
import locale
locale_str = 'fr_FR'

file = 'file_example_XLSX_50.xlsx'


my_obj_names = ['url', 'imgUrl', 'title', 'desc', 'date']
first_row = 2
last_row = 14  # exclusive
first_col = 2
last_col = 7  # exclusive

date_format = '%d %b %Y'


# Don't touch the code after this line
locale.setlocale(locale.LC_ALL, locale_str)
locale.getpreferredencoding('utf-8')
wb = load_workbook(f'xlsx/{file}')
ws = wb.active

if len(my_obj_names) != last_col - first_col:
    print("The names number doesn't fit the col number")
    raise SystemExit

finalArray = []

for row in range(first_row, last_row):
    dic = {}
    for col in range(first_col, last_col):
        char = get_column_letter(col)
        value = ws[char + str(row)].value

        if type(value) is datetime:
            value = value.strftime(date_format)

        dic[my_obj_names[col - 2]] = value

    finalArray.append(dic)

with open(f'json/{file.split(".")[0]}.json', 'w') as f:
    json.dump(finalArray, f, indent=2)
    print(f"New json file is created from {file} file")
